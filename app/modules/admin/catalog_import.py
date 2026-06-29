from fastapi import APIRouter, Depends, HTTPException
from sqlalchemy.orm import Session
from sqlalchemy import text
import boto3
import io
from decimal import Decimal
from uuid import uuid4
from openpyxl import load_workbook

from app.database import get_db
from app.modules.auth.dependencies import require_admin
from app.modules.auth.models import User
from app.modules.catalog.models import ProductCategory, Brand, Product, ProductVariant
from app.config import settings

router_admin = APIRouter(prefix="/admin/catalog", tags=["Admin — Catalogo"])

S3_IMPORT_KEY = "imports/productos.xlsx"

# Campos que viven solo en la fila template (variante=Unico, agregar=no) y que
# las filas de variantes heredan automaticamente del producto ya creado.
TEMPLATE_FIELDS = [
    "sku_template", "retail_price", "descripcion", "modo_de_uso",
    "beneficios", "ingredientes", "atributos", "etiquetas",
]


def _get_worksheet():
    if settings.environment == "local":
        wb = load_workbook("productos.xlsx", data_only=True)
    else:
        s3 = boto3.client("s3", region_name=settings.aws_region)
        response = s3.get_object(Bucket=settings.s3_bucket_name, Key=S3_IMPORT_KEY)
        content = response["Body"].read()
        wb = load_workbook(io.BytesIO(content), data_only=True)
    return wb.active


def _notna(value) -> bool:
    return value is not None and str(value).strip() != ""


def parse_atributos(raw):
    if not raw:
        return None
    result = {}
    for pair in str(raw).split("|"):
        pair = pair.strip()
        if ":" in pair:
            key, value = pair.split(":", 1)
            result[key.strip()] = value.strip()
    return result if result else None


def _read_template_fields(row: dict) -> dict:
    """Lee del row los campos que solo existen en la fila template."""
    descripcion = str(row["descripcion"]).strip() if _notna(row.get("descripcion")) else None
    etiquetas_raw = str(row["etiquetas"]).strip() if _notna(row.get("etiquetas")) else None
    if etiquetas_raw and etiquetas_raw.lower() == "ninguno":
        etiquetas = []
    elif etiquetas_raw:
        etiquetas = [t.strip() for t in etiquetas_raw.split(",")]
    else:
        etiquetas = None

    precio_menudeo = Decimal(str(row["precio_menudeo"])) if _notna(row.get("precio_menudeo")) else None
    precio_mayoreo = Decimal(str(row["precio_mayoreo"])) if _notna(row.get("precio_mayoreo")) else None
    margen_pct = Decimal(str(row["margen_mayoreo_pct"])) if _notna(row.get("margen_mayoreo_pct")) else None
    cost_price = round(precio_mayoreo * (1 - margen_pct / 100), 2) if (precio_mayoreo is not None and margen_pct is not None) else None

    return {
        "retail_price": precio_menudeo,
        "list_price": precio_mayoreo,
        "cost_price": Decimal(str(cost_price)) if cost_price is not None else None,
        "descripcion": descripcion,
        "modo_de_uso": str(row["modo_de_uso"]).strip() if _notna(row.get("modo_de_uso")) else None,
        "beneficios": str(row["beneficios"]).strip() if _notna(row.get("beneficios")) else None,
        "ingredientes": str(row["ingredientes"]).strip() if _notna(row.get("ingredientes")) else None,
        "atributos": parse_atributos(row.get("atributos")),
        "etiquetas": etiquetas,
    }


def _apply_template_fields(producto: Product, fields: dict) -> None:
    if fields.get("sku_template") is not None:
        producto.sku_template = fields["sku_template"]
        # Imagen general del producto (fallback) — siempre mayusculas, igual
        # al SKU real. Se deriva de sku_template, nunca del sku de una
        # variante especifica.
        bucket = settings.s3_bucket_name
        sku_tpl = fields["sku_template"]
        producto.image_url = f"https://{bucket}.s3.amazonaws.com/productos/{sku_tpl}.jpg"
        producto.image_thumb_url = producto.image_url
    if fields.get("retail_price") is not None:
        producto.retail_price = fields["retail_price"]
    if fields.get("list_price") is not None:
        producto.list_price = fields["list_price"]
    if fields.get("cost_price") is not None:
        producto.cost_price = fields["cost_price"]
    if fields.get("descripcion") is not None:
        producto.description = fields["descripcion"]
    if fields.get("modo_de_uso") is not None:
        producto.modo_de_uso = fields["modo_de_uso"]
    if fields.get("beneficios") is not None:
        producto.beneficios = fields["beneficios"]
    if fields.get("ingredientes") is not None:
        producto.ingredientes = fields["ingredientes"]
    if fields.get("atributos") is not None:
        producto.atributos = fields["atributos"]
    if fields.get("etiquetas") is not None:
        producto.tags = fields["etiquetas"]


def _ensure_categoria_marca(db: Session, row: dict):
    """
    Crea o actualiza categoria y marca a partir de cualquier fila — template
    o variante. Se llama tanto en la fila template (que trae margen_mayoreo_pct
    real) como al inicio de _process_row() para filas de variante (donde ese
    campo viene en blanco por diseno y no debe pisar el valor ya guardado).
    Sin esto, una marca/categoria nueva nunca se creaba si el catalogo estaba
    vacio: las filas template se interceptan con continue en import_catalog()
    antes de llegar a _process_row(), asi que esta logica nunca se ejecutaba
    para ellas — la primera fila que de verdad la alcanzaba era una variante
    con margen_mayoreo_pct en blanco, lo cual fallaba al crear una marca nueva.
    """
    categoria_nombre = str(row["categoria"]).strip()
    marca_nombre = str(row["marca"]).strip()
    margen_pct = Decimal(str(row["margen_mayoreo_pct"])) if _notna(row.get("margen_mayoreo_pct")) else None

    categoria = db.query(ProductCategory).filter(ProductCategory.name == categoria_nombre).first()
    if not categoria:
        slug_base = categoria_nombre.lower().replace(" ", "-").replace("á","a").replace("é","e").replace("í","i").replace("ó","o").replace("ú","u")
        # Verificar que el slug no exista ya
        slug = slug_base
        slug_exists = db.query(ProductCategory).filter(ProductCategory.slug == slug).first()
        if slug_exists:
            slug = f"{slug_base}-{str(uuid4())[:4]}"
        categoria = ProductCategory(id=uuid4(), name=categoria_nombre, slug=slug, active=True, display_order=0)
        db.add(categoria)
        db.flush()

    marca = db.query(Brand).filter(Brand.name == marca_nombre).first()
    if not marca:
        if margen_pct is None:
            raise ValueError(
                f"No se puede crear la marca '{marca_nombre}' sin margen_mayoreo_pct "
                f"— falta la fila template con ese dato"
            )
        marca = Brand(id=uuid4(), name=marca_nombre, brand_discount_percentage=margen_pct, active=True)
        db.add(marca)
        db.flush()
    elif margen_pct is not None:
        marca.brand_discount_percentage = margen_pct

    return categoria, marca


def _process_row(db: Session, row: dict) -> dict:
    # margen_mayoreo_pct y precio_mayoreo son "solo template" en el formato
    # nuevo — en filas de variante vienen en blanco, se heredan del producto
    # ya creado por la fila template (que el import siempre procesa primero).
    margen_pct = Decimal(str(row["margen_mayoreo_pct"])) if _notna(row.get("margen_mayoreo_pct")) else None
    producto_nombre = str(row["nombre"]).strip()
    precio_mayoreo = Decimal(str(row["precio_mayoreo"])) if _notna(row.get("precio_mayoreo")) else None
    sku = str(row["sku"]).strip()
    variante_nombre = str(row["variante"]).strip()
    agregar = str(row["agregar"]).strip().lower()
    # Convencion de galeria (3B): primera imagen de la variante = {SKU}.jpg
    # (mayusculas, sin sufijo — el sufijo _2, _3... es solo para imagenes
    # adicionales subidas despues, fuera del flujo de import del Excel).
    image_url = f"https://{settings.s3_bucket_name}.s3.amazonaws.com/productos/{sku}.jpg"

    categoria, marca = _ensure_categoria_marca(db, row)
    if agregar == "no":
        # Si es el template (variante=Unico) guardar campos template sin crear variante
        if variante_nombre == "Unico":
            # La marca y categoria ya fueron procesadas arriba
            producto = db.query(Product).filter(
                Product.name == producto_nombre,
                Product.brand_id == marca.id,
            ).first()
            if producto:
                template_fields = _read_template_fields(row)
                template_fields["sku_template"] = sku
                _apply_template_fields(producto, template_fields)
                return {"action": "updated", "sku": sku}
            return {"action": "not_found", "sku": sku}
        # Si no es template — eliminar variante por SKU
        variante = db.query(ProductVariant).filter(ProductVariant.sku == sku).first()
        if variante:
            otras_variantes = db.query(ProductVariant).filter(
                ProductVariant.product_id == variante.product_id,
                ProductVariant.id != variante.id,
            ).count()
            db.delete(variante)
            if otras_variantes == 0:
                producto = db.query(Product).filter(Product.id == variante.product_id).first()
                if producto:
                    db.delete(producto)
            return {"action": "deleted", "sku": sku}
        return {"action": "not_found", "sku": sku}


    # precio_mayoreo/margen_pct solo vienen en la fila template — en filas de
    # variante (precio_mayoreo is None) se hereda list_price/cost_price del
    # producto ya creado, en vez de recalcular con datos en blanco.
    cost_price = round(precio_mayoreo * (1 - margen_pct / 100), 2) if (precio_mayoreo is not None and margen_pct is not None) else None
    slug_producto = producto_nombre.lower().replace(" ", "-").replace("á","a").replace("é","e").replace("í","i").replace("ó","o").replace("ú","u")
    variante_display = None if variante_nombre == "Unico" else variante_nombre
    variante = db.query(ProductVariant).filter(ProductVariant.sku == sku).first()

    if variante:
        producto = db.query(Product).filter(Product.id == variante.product_id).first()
        if producto:
            producto.name = producto_nombre
            if precio_mayoreo is not None:
                producto.list_price = precio_mayoreo
                producto.cost_price = Decimal(str(cost_price))
            producto.category_id = categoria.id
            producto.brand_id = marca.id
            producto.slug = f"{slug_producto}-{str(producto.id)[:8]}"
        variante.variant_name = variante_display
        action = "updated"
    else:
        producto = db.query(Product).filter(Product.name == producto_nombre, Product.brand_id == marca.id).first()
        if not producto:
            # precio_mayoreo puede venir en blanco aqui si esta fila es la
            # primera variante de un producto nuevo cuyo precio vive en la
            # fila template — import_catalog() aplica campos_template_pendientes
            # sobre este mismo producto inmediatamente despues de este return,
            # en la misma iteracion, antes del commit final. Placeholder temporal.
            producto = Product(
                id=uuid4(), name=producto_nombre,
                slug=f"{slug_producto}-{str(uuid4())[:8]}",
                category_id=categoria.id, brand_id=marca.id,
                list_price=precio_mayoreo if precio_mayoreo is not None else Decimal("0.00"),
                cost_price=Decimal(str(cost_price)) if cost_price is not None else Decimal("0.00"),
                # image_url/image_thumb_url del producto se derivan de sku_template
                # via _apply_template_fields() — no del sku de esta fila de variante.
                active=True, display_order=0,
            )
            db.add(producto)
            db.flush()
            action = "created"
        else:
            if precio_mayoreo is not None:
                producto.list_price = precio_mayoreo
                producto.cost_price = Decimal(str(cost_price))
            action = "updated"

        variante = ProductVariant(
            id=uuid4(), product_id=producto.id, sku=sku,
            variant_name=variante_display, stock_qty=0, returned_stock_qty=0,
            image_url=image_url, active=True, display_order=0,
        )
        db.add(variante)
        if action == "updated":
            action = "variant_added"

    # Si es producto simple (variante=Unico y agregar=si) guarda los campos template
    # directamente en esta misma fila — no hay fila template separada.
    if variante_nombre == "Unico" and producto:
        template_fields = _read_template_fields(row)
        template_fields["sku_template"] = sku
        _apply_template_fields(producto, template_fields)

    return {"action": action, "sku": sku}


@router_admin.post("/import")
def import_catalog(
    current_user: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    try:
        ws = _get_worksheet()
    except Exception as e:
        raise HTTPException(status_code=422, detail=f"No se pudo leer el archivo: {str(e)}")

    headers = [cell.value for cell in ws[1]]
    required = {"categoria", "marca", "margen_mayoreo_pct", "nombre", "precio_mayoreo", "sku", "variante", "agregar"}
    missing = required - set(h for h in headers if h)
    if missing:
        raise HTTPException(status_code=422, detail=f"Columnas faltantes en el Excel: {missing}")

    created = updated = deleted = not_found = errors = 0
    nombre_template_pendiente = None
    campos_template_pendientes = None

    for idx, row_values in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(v is not None for v in row_values):
            continue
        row = dict(zip(headers, row_values))
        agregar_val = str(row.get('agregar', '')).strip().lower()
        variante_val = str(row.get('variante', '')).strip()
        nombre_val = str(row.get('nombre', '')).strip()
        if variante_val == 'Unico' and agregar_val == 'no':
            try:
                _ensure_categoria_marca(db, row)
            except Exception as e:
                errors += 1
                print(f"ERROR fila {idx}: {str(e)}")
                db.rollback()
                continue
            campos_template_pendientes = _read_template_fields(row)
            campos_template_pendientes["sku_template"] = str(row.get('sku', '')).strip()
            nombre_template_pendiente = nombre_val
            continue
        try:
            result = _process_row(db, row)
            if campos_template_pendientes and nombre_template_pendiente == nombre_val:
                prod_obj = db.query(Product).filter(Product.name == nombre_val).first()
                if prod_obj:
                    _apply_template_fields(prod_obj, campos_template_pendientes)
                    campos_template_pendientes = None
                    nombre_template_pendiente = None
            if result["action"] == "created": created += 1
            elif result["action"] in ("updated", "variant_added"): updated += 1
            elif result["action"] == "deleted": deleted += 1
            elif result["action"] == "not_found": not_found += 1
        except Exception as e:
            errors += 1
            print(f"ERROR fila {idx}: {str(e)}")
            db.rollback()
            continue

    db.commit()
    return {"status": "completed", "created": created, "updated": updated, "deleted": deleted, "not_found": not_found, "errors": errors}


@router_admin.delete("/clear")
def clear_catalog(
    current_user: User = Depends(require_admin),
    db: Session = Depends(get_db),
):
    try:
        db.execute(text("SET session_replication_role = replica"))
        db.execute(text("DELETE FROM product_images"))
        db.execute(text("DELETE FROM product_variants"))
        db.execute(text("DELETE FROM products"))
        db.execute(text("DELETE FROM product_categories"))
        db.execute(text("DELETE FROM brands"))
        db.execute(text("SET session_replication_role = DEFAULT"))
        db.commit()
        return {"status": "completed", "message": "Catalogo eliminado correctamente"}
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))
